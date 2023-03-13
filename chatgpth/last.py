import openpyxl
import socket
import datetime

DEFAULT_URL = "https://www.example.com"

# URL 주소를 IP 주소로 변환하는 함수
def url_to_ip(url):
    try:
        ip = socket.gethostbyname(url)
        return ip
    except:
        return None

# 엑셀 파일을 열고 시트를 생성하는 함수
def create_excel():
    wb = openpyxl.Workbook()
    org_data = wb.active
    org_data.title = "org_data"
    org_data.append(["Date", "IP"])
    compare_data = wb.create_sheet("compare_data")
    compare_data.append(["Date", "IP"])
    return wb, org_data, compare_data

# IP 주소를 저장하는 함수
def save_ip(ip, sheet):
    date = datetime.datetime.today().strftime("%Y-%m-%d")
    if sheet.title == "org_data":
        sheet.append([date, ip])
    elif sheet.title == "compare_data":
        last_row = sheet.max_row
        last_ip = sheet.cell(row=last_row, column=2).value
        if last_ip != ip:
            sheet.append([date, ip])

# URL을 입력받고 IP 주소를 저장하는 함수
def url_to_ip_excel():
    url = input("Enter URL: ")
    if not url:
        url = DEFAULT_URL
        print(f"Using default URL: {DEFAULT_URL}")
    wb_exists = openpyxl.load_workbook("ip_data.xlsx")
    org_data = wb_exists["org_data"]
    compare_data = wb_exists["compare_data"]
    ip = url_to_ip(url)
    if ip:
        save_ip(ip, org_data)
        save_ip(ip, compare_data)
        wb_exists.save("ip_data.xlsx")
        print("IP address saved successfully.")
    else:
        print("Failed to convert URL to IP address.")