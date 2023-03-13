import socket
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# 오늘 날짜 정보 생성
today = date.today()
today_str = today.strftime("%Y-%m-%d")

# 엑셀 파일 로드
try:
    wb = load_workbook('example.xlsx')
except FileNotFoundError:
    # 엑셀 파일이 없는 경우, 새로운 파일 생성
    wb = Workbook()
    ws_url = wb.active
    ws_url.title = 'URL'
    ws_url.append(['URL'])
    ws1 = wb.create_sheet('org_data')
    ws1.append(['Date', 'URL', 'IP Address'])
    ws1.freeze_panes = ws1['A2']
    ws2 = wb.create_sheet('compare_data')
    ws2.append(['Date', 'URL', 'Previous IP', 'Current IP'])
    ws2.freeze_panes = ws2['A2']

    # 테이블 스타일 적용
    for ws in [ws_url, ws1, ws2]:
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.row_dimensions[1].height = 20
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            cell.fill = PatternFill(patternType='solid', fgColor='C6E0B4')
            cell.border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                  top=Side(style='medium'), bottom=Side(style='medium'))

else:
    # 엑셀 파일이 있는 경우, 시트 가져오기
    ws_url = wb['URL']
    ws1 = wb['org_data']
    ws2 = wb['compare_data']

# URL 리스트 가져오기
url_list = []
for row in ws_url.iter_rows(min_row=2, values_only=True):
    url_list.append(row[0])

# 첫번째 시트에 데이터 입력
for url in url_list:
    ip_address = socket.gethostbyname(url)
    ws1.append([today_str, url, ip_address])

# 두번째 시트에 데이터 입력
if len(ws1['A']) > 1:
    prev_date = ws1['A'][-2].value
    prev_ips = {}
    for row in ws1.iter_rows(min_row=2, values_only=True):
        date, url, ip_address = row
        if date == prev_date:
            if prev_ips.get(url) != ip_address:
                ws2.append([date, url, prev_ips.get(url), ip_address])
        else:
            prev_date = date
            prev_ips = {}
        prev_ips[url] = ip_address

# 엑셀 파일 저장
wb.save('example.xlsx')